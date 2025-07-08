
import pathlib
import zipfile
from datetime import datetime
import os
import json
import re
import time
from memorious.operations.extract import extract_zip
from openpyxl import load_workbook, Workbook
from requests.exceptions import RequestException
import lxml.html

from concurrent.futures import ThreadPoolExecutor
# Import from ua_kmr_voting_excel_functions.py function find_last_row
from ua_kmr_voting_excel_functions import find_last_row, request_page, count_non_empty_columns, transform_results, preprocess_json


import time
timestamp = int(time.time())  # Add a timestamp for uniqueness

# Counter to ensure uniqueness
counter = 0

# Adjust this as needed, depending on the number of threads you want to use
MAX_WORKERS = 4
zipfile_paths = []
CHUNK_SIZE = 100 
URL_FULL = "https://kmr.gov.ua/uk/result_golosuvanya?title=&field_start_date_n_h_value%5Bmin%5D=&field_start_date_n_h_value%5Bmax%5D=&page="
URL = 'https://kmr.gov.ua/uk/result_golosuvanya?title=&field_start_date_n_h_value[min]&field_start_date_n_h_value[max]&page=0&start=0'
HEADERS_XLSX = ["id",  "time", "question", "status", "short_name", "result", "file_name", "source_url"]



### PROCESS XLSX FILES ###
processed_files = set()

def process_xlsx_file(context, path):
    """Process XLSX files function."""
    new_path = path.with_suffix('.xlsx')
    print(f"Processing XLSX file: {path} -> {new_path}")

    if new_path in processed_files:
        context.log.info(f"Skipping already processed file: {new_path}")
        return

    with open(new_path, 'rb') as f:
        file_content = f.read()

    with open(new_path, 'wb') as f:
        f.write(file_content)

    processed_files.add(new_path)

    # Process the XLSX file
    process_xlsx(context, new_path)


import requests
def download_and_process(context, link, is_zip):
    """Download and process files in parallel."""
    print("Downloading and processing files in parallel.")
    #response = request_page(context, link)
    print(f"Downloading file from: {link}")
    response = requests.get(link)

   

    if response:

        #print(response.headers)  # Headers of the response

        file_name = link.split('/')[-1]
        print(f"File name in response: {file_name}")  # File name: 14.05.22_1_z_kovalchukom.zip
    
        file_path = pathlib.Path(context.work_path) / file_name  # Create a file path in the work directory
        print(f"File path in response: {file_path}")  # File path: /
        #file_path = pathlib.Path(response.file_path)
        #print(f"File path in response: {file_path}")  # File path: /var/folders/gx/61q8sz7d4r9g0nl993whgmhc0000gp/T/tmpsdaxn8pb/14.05.22_1_z_kovalchukom.zip
        
        ### Process ZIP files. If True, process ZIP files.
        if is_zip:  
            # add a suffix to the file name
            file_path = file_path.with_suffix('.zip')
            os.rename(response.file_path, file_path)
 
            extract_to = pathlib.Path(context.work_path) / file_name # create a directory to extract into
            context.log.info(f"Downloaded and stored ZIP file: {file_name}, in {extract_to}") # Downloaded and stored ZIP file: 14.05.22_1_z_kovalchukom.zip, in /var/folders/gx/61q8sz7d4r9g0nl993whgmhc0000gp/T/tmpsdaxn8pb/14.05.22_1_z_kovalchukom.zip 
            #zipfile_paths.append(file_path)

            # Unzip the file and process the extracted files
            # with zipfile.ZipFile(file_path, "r") as zf:
            #     zf.extractall(extract_to)
            # context.log.info(f"Extracted ZIP file: {file_path}")

            # Process the extracted files
            #process_extracted_files(context, extract_to, link)

        ### Process XLSX files. If False, process XLSX files.
        if not is_zip:  
            # add a suffix to the file name
            file_path = file_path.with_suffix('.xlsx')
            #os.rename(response.file_path, file_path)
            context.log.info(f"Downloaded and stored XLSX file: {file_name}, in {file_path}")
            # Process the XLSX file
            # SAVE THE DOWNLOADED CONTENT TO DISK
            try:
                with open(file_path, 'wb') as f:
                    f.write(response.content)
            except Exception as e:
                context.log.error(f"Failed to save XLSX file {file_path}: {e}")
                return
                
            process_xlsx_file(context, file_path)
        # else:  
        #     process_xlsx_file(context, file_path)
    else:
        context.log.error(f"Failed to download file: {link}")



def process_extracted_files(context, extract_to, link):

    context.log.info(f"Processing extracted files in the directory: {extract_to}")

    extracted_files = list(pathlib.Path(extract_to).iterdir())  # Find all extracted files in the directory

    # Fist all files in the directory extracted_files
    for extracted_file in extracted_files[:2]: # DELETE [:2] LATER

        # List files in the extracted file
        for json_file in extracted_file.iterdir():

            # Preprocess and read the JSON file
            corrected_json = preprocess_json(json_file)

            # Parse the JSON data
            try:
                json_data = json.loads(corrected_json)
            except json.JSONDecodeError as e:
                print(f"Error decoding JSON: {e}")
                return


            # Extract and store data in the dict
            dict_d = {
                'file_name': str(json_file),
                'file_path': str(extract_to),
                'extracted_file': str(extracted_file),
                'source_url': link,

                'DocTime': json_data['DocTime'],
                'orgName': json_data['OrgName'],
                'SName': json_data['SName'],
                'GLType': json_data['GLType'],
                'GLTime': json_data['GLTime'],
                #'Date_GLTime': datetime.strptime(json_data['GLTime'], '%d.%m.%Y').isoformat(),
                #'Time_GLTime': datetime.strptime(json_data['GLTime'], '%H:%M:%S').isoformat(),
                'PD_NPP': json_data['PD_NPP'],
                'GL_Text': json_data['GL_Text'],
                'DPList': json_data['DPList'],
                'RESULT': json_data['RESULT']
            }
            #store_json(context, dict_d)


### MAIN FUNCTION ###
def init(context, data):
    for i in range(1, 18): # XLSX and JSON files are in range(1, 18)
    #for i in range(10, 11): # Adjust range as needed. range(16, 17) contains XLSX files.
    #for i in range(1, 2):
        URL_FULL_PAGE = URL_FULL + str(i)
        response_main_page = request_page(context, URL_FULL_PAGE)
        doc = lxml.html.fromstring(response_main_page.text)
        block = doc.xpath('//div[@class="view-content"]')

        for item in block[:1]:
            for j in item.getchildren():

                if 'zip' in j.text_content():  # ZIP files
                    continue # DELETE LATER
                    # link = j.xpath('.//a/@href')[1]
                    # print("\n")
                    # context.log.info(f"Processing ZIP file: {link}")
                    #download_and_process(context, link, True) # Download and process ZIP files

                elif 'xlsx' in j.text_content():  # XLSX files
                    link_xlsx = j.xpath('.//a/@href')[1]
                    context.log.info(f"Processing XLSX file: {link_xlsx}")
                    download_and_process(context, link_xlsx, False)




def process_xlsx(context, file_path):
    """Process the XLSX files and emit rows."""
    context.log.info(f"Processing XLSX file in Process XLSX function: {file_path}")
    xlsx_files = [file for file in os.listdir(context.work_path) if file.endswith('.xlsx')]
    context.log.info(f"Found {len(xlsx_files)} XLSX files in the work path: {context.work_path}")

    for xlsx_file in xlsx_files:
        context.log.info(f"Processing XLSX file in the loop: {xlsx_file}")

        file_path = os.path.join(context.work_path, xlsx_file)
        print(f"File path in the loop: {file_path}")  # File path: /var/folders/gx/61q8sz7d4r9g0nl993whgmhc0000gp/T/tmpsdaxn8pb/14.05.22_1_z_kovalchukom.xlsx
        wb = load_workbook(file_path, read_only=False)
        print(f"Workbook loaded: {wb.sheetnames}")  # Debugging line to see the sheet names
        ws = wb[wb.sheetnames[0]]
        print(f"Worksheet selected: {ws.title}")  # Debugging line to see the selected worksheet

        max_row_N = find_last_row(ws)
        non_empty_col_count = count_non_empty_columns(ws) # Count non-empty columns in the first row
        
        start_col = 5 if non_empty_col_count == 125 else 6 # Adjust start column based on the number of columns

        fixed_cols_range = range(1, start_col)

        for row in ws.iter_rows(min_row=2, max_row=max_row_N, values_only=True):
            fixed_data = {f"fixed_col_{col}": row[col - 1] for col in fixed_cols_range}
            for col in range(start_col, start_col + 122):  # Adjust end column based on the number of columns
                entity_name = ws.cell(row=1, column=col).value
                result = row[col - 1] 

                # Create a dictionary for the row
                new_row = {
                    **fixed_data,
                    "entity_name": entity_name,
                    "result": result,
                    "xlsx_file": xlsx_file # hash
                }
                print(f"New row data: {new_row}")  # Debugging line to see the new row data
                #context.emit(data=new_row)
                #store_excel(context, new_row)


def store_excel(context, data):
    context.log.info(f"Storing XLSX data from: {data.get('xlsx_file')}")
    
    # Document data
    doc_data = {
        'id_original': data.get('fixed_col_1'),  # Original ID for reference
        'DocTime': str(data.get('fixed_col_2')), 
        'GL_Text': data.get('fixed_col_3'), 
        'RESULT': data.get('fixed_col_4'),  
        'DPName': data.get('entity_name'),  
        'DPGolos': data.get('result'), 
        "file_name": data.get('xlsx_file'),
        'retrieved_at': datetime.now().isoformat()
    }
    
    # Establish connection to SQLite database
    
    conn = sqlite3.connect('voting_excel_data.db')
    cursor = conn.cursor()

    # Ensure the table exists (CREATE IF NOT EXISTS)
    create_table_sql = """
    CREATE TABLE IF NOT EXISTS ua_kmr_voting_xlsx (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_original TEXT NOT NULL,
        DocTime TEXT,
        GL_Text TEXT,
        RESULT TEXT,
        DPName TEXT,
        DPGolos TEXT,
        file_name TEXT,
        retrieved_at TEXT,
        source_url TEXT,
        UNIQUE(id_original, DocTime, GL_Text, RESULT, DPName, DPGolos, file_name)
    );
    """
    cursor.execute(create_table_sql)
    
    # Check if the record exists based on the unique fields (you may modify the unique check as needed)
    check_record_sql = """
    SELECT * FROM ua_kmr_voting_xlsx WHERE id_original = ? AND DocTime = ? AND GL_Text = ? 
          AND RESULT = ? AND DPName = ? AND DPGolos = ? AND file_name = ?;
    """
    cursor.execute(check_record_sql, (doc_data['id_original'], doc_data['DocTime'], doc_data['GL_Text'], 
                                      doc_data['RESULT'], doc_data['DPName'], doc_data['DPGolos'], doc_data['file_name']))
    existing_record = cursor.fetchone()
    
    if existing_record:
        # If record exists, update it
        update_sql = """
        UPDATE ua_kmr_voting_xlsx
        SET DocTime = ?, GL_Text = ?, RESULT = ?, DPName = ?, DPGolos = ?, retrieved_at = ?
        WHERE id_original = ? AND file_name = ?;
        """
        cursor.execute(update_sql, (
            doc_data['DocTime'], doc_data['GL_Text'], doc_data['RESULT'], doc_data['DPName'], doc_data['DPGolos'], 
            doc_data['retrieved_at'], doc_data['id_original'], doc_data['file_name']
        ))
        context.log.info(f"Updated existing record with id_original: {doc_data['id_original']}")
    else:
        # If record doesn't exist, insert a new one
        insert_sql = """
        INSERT INTO ua_kmr_voting_xlsx (id_original, DocTime, GL_Text, RESULT, DPName, DPGolos, file_name, retrieved_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?);
        """
        cursor.execute(insert_sql, (
            doc_data['id_original'], doc_data['DocTime'], doc_data['GL_Text'], doc_data['RESULT'], 
            doc_data['DPName'], doc_data['DPGolos'], doc_data['file_name'], doc_data['retrieved_at']
        ))
        context.log.info(f"Inserted new record with id_original: {doc_data['id_original']}")
    
    # Commit the transaction and close the connection
    conn.commit()
    conn.close()

    context.log.info("Data stored successfully.")


### STORE JSON DATA ###

# import sqlite3
# from datetime import datetime

# def store_json(context, data):
#     # Log the data processing
#     context.log.info(f"Storing JSON data from: {data.get('file_name')}")

#     # Prepare the main document data
#     doc_data = {
#         'id': data.get('id'),
#         'file_name': data.get('file_name').split('/')[-1],
#         'Num_Question': data.get('file_name').split('/')[-1].split('_')[1].split('.')[0],
#         'source_url': data.get('source_url'),
#         'orgName': "КИЇВСЬКА МІСЬКА РАДА",
#         'SName': data.get('SName'),
#         'GLType': data.get('GLType'),
#         'GLTime': data.get('GLTime'),
#         'PD_NPP': data.get('PD_NPP'),
#         'GL_Text': data.get('GL_Text'),
#         'DocTime': data.get('DocTime'),
#         'RESULT': data.get('RESULT'),
#         'retrieved_at': datetime.now().isoformat()
#     }

#     # Prepare the combined data
#     combined_data = []
#     dp_list = data.get('DPList', [])
#     for dp_entry in dp_list:
#         combined_entry = doc_data.copy()
#         combined_entry.update({
#             'DPName': dp_entry.get('DPName'),
#             'DPGolos': 'Відсутній' if dp_entry.get('DPGolos') == '.........' else dp_entry.get('DPGolos')
#         })
#         combined_data.append(combined_entry)
    
#     if not dp_list:
#         combined_data.append(doc_data)
    
#     # Connect to SQLite database
#     conn = sqlite3.connect('voting_json_data.db')
#     cursor = conn.cursor()

#     # Create the table if it doesn't exist
#     cursor.execute('''
#         CREATE TABLE IF NOT EXISTS ua_kmr_voting_json (
#             id TEXT,
#             file_name TEXT,
#             Num_Question TEXT,
#             source_url TEXT,
#             orgName TEXT,
#             SName TEXT,
#             GLType TEXT,
#             GLTime TEXT,
#             PD_NPP TEXT,
#             GL_Text TEXT,
#             DocTime TEXT,
#             DPName TEXT,
#             DPGolos TEXT,
#             RESULT TEXT,
#             retrieved_at TEXT,
#             UNIQUE(id, file_name, Num_Question, SName, GLType, GLTime, PD_NPP, GL_Text, DocTime, DPName, RESULT)
#         )
#     ''')

#     # Insert or update each entry
#     for entry in combined_data:
#         cursor.execute('''
#             INSERT INTO ua_kmr_voting_json (id, file_name, Num_Question, source_url, orgName, SName, GLType, 
#                                              GLTime, PD_NPP, GL_Text, DocTime, DPName, DPGolos, RESULT, retrieved_at)
#             VALUES (:id, :file_name, :Num_Question, :source_url, :orgName, :SName, :GLType, 
#                     :GLTime, :PD_NPP, :GL_Text, :DocTime, :DPName, :DPGolos, :RESULT, :retrieved_at)
#             ON CONFLICT(id, file_name, Num_Question, SName, GLType, GLTime, PD_NPP, GL_Text, DocTime, DPName, RESULT)
#             DO UPDATE SET
#                 source_url = excluded.source_url,
#                 orgName = excluded.orgName,
#                 SName = excluded.SName,
#                 GLType = excluded.GLType,
#                 GLTime = excluded.GLTime,
#                 PD_NPP = excluded.PD_NPP,
#                 GL_Text = excluded.GL_Text,
#                 DocTime = excluded.DocTime,
#                 DPName = excluded.DPName,
#                 DPGolos = excluded.DPGolos,
#                 RESULT = excluded.RESULT,
#                 retrieved_at = excluded.retrieved_at
#         ''', entry)

#     # Commit changes and close the connection
#     conn.commit()
#     conn.close()


# def store_json(context, data):
#     # Extract main document data
#     context.log.info(f"Storing JSON data from: {data.get('file_name')}")

#     #'Date_GLTime': datetime.strptime(json_data['GLTime'], '%d.%m.%Y').isoformat(),
#     #'Time_GLTime': datetime.strptime(json_data['GLTime'], '%H:%M:%S').isoformat(),
#     doc_data = {
#         'id': data.get('id'),
           
#         'file_name': data.get('file_name').split('/')[-1],
#         'Num_Question': data.get('file_name').split('/')[-1].split('_')[1].split('.')[0],

#         'source_url': data.get('source_url'),
#         'orgName': "КИЇВСЬКА МІСЬКА РАДА",
#         'SName': data.get('SName'),
#         'GLType': data.get('GLType'),

#         'GLTime': data.get('GLTime'),
#         # 'Date_GLTime': data.get('GLTime').split("")[0],
#         # 'Time_GLTime': data.get('Time_GLTime').split("")[1],

#         'PD_NPP': data.get('PD_NPP'),
#         'GL_Text': data.get('GL_Text'),
#         'DocTime': data.get('DocTime'),
#         'RESULT': data.get('RESULT'),
#         'retrieved_at': datetime.now().isoformat()
#     }

#     # Process and combine DPList entries with main document data
#     combined_data = []
#     dp_list = data.get('DPList', [])
#     for dp_entry in dp_list:
#         combined_entry = doc_data.copy()
#         combined_entry.update({
#             'DPName': dp_entry.get('DPName'),
#             'DPGolos': 'Відсутній' if dp_entry.get('DPGolos') == '.........' else dp_entry.get('DPGolos')
#         })
#         #print("Combined entry:", combined_entry)
#         combined_data.append(combined_entry)
    
#     if not dp_list:
#         combined_data.append(doc_data)
    
#     table = context.datastore['ua_kmr_voting_json']
#     for entry in combined_data:

#         table.upsert(entry, ['id', 'file_name', 'Num_Question', 'SName', 'GLType', 
#                              'GLTime', 'Date_GLTime', 'Time_GLTime',
#                              'PD_NPP', 'GL_Text', 
#                              'DocTime', 'DPName', 'DPGolos', 'RESULT', 'retrieved_at'])




import sqlite3
from datetime import datetime







# def store_exel(context, data):

#     context.log.info(f"Storing XLSX data from: {data.get('xlsx_file')}")

#     doc_data = {
#         #'id': unique_id,  # Ensure id is always a string
#         'id_original': data.get('fixed_col_1'), # Keep original ID for reference

#         'DocTime': str(data.get('fixed_col_2')), 
#         'GL_Text': data.get('fixed_col_3'), 
#         'RESULT': data.get('fixed_col_4'),  
#         'DPName': data.get('entity_name'),  
#         'DPGolos': data.get('result'), 
#         "file_name": data.get('xlsx_file'),
#         'retrieved_at': datetime.now().isoformat()
#     }

#     table = context.datastore['ua_kmr_voting_xlsx']
#     table.upsert(doc_data, [#'id', # id is present by default, no need to specify
#                             'id_original', 'DocTime',
#                             'GL_Text', 'RESULT', 'DPName', 'DPGolos',
#                             'file_name','retrieved_at', 'source_url'])

