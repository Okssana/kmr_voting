import time
from requests.exceptions import RequestException
from openpyxl import Workbook, load_workbook

# Counter to ensure uniqueness
counter = 0

# Adjust this as needed, depending on the number of threads you want to use
MAX_WORKERS = 4
zipfile_paths = []
CHUNK_SIZE = 100 
URL_FULL = "https://kmr.gov.ua/uk/result_golosuvanya?title=&field_start_date_n_h_value%5Bmin%5D=&field_start_date_n_h_value%5Bmax%5D=&page="
URL = 'https://kmr.gov.ua/uk/result_golosuvanya?title=&field_start_date_n_h_value[min]&field_start_date_n_h_value[max]&page=0&start=0'
HEADERS_XLSX = ["id",  "time", "question", "status", "short_name", "result", "file_name", "source_url"]


import re
# Function to preprocess JSON data
def preprocess_json(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Correct invalid newline characters within strings
    corrected_content = re.sub(r'(?<!\\)"\s*\n\s*"', r'\"', content)
    
    return corrected_content

def request_page(context, url):

    for attempt in range(1, 5):
        try:
            response = context.http.get(url)
            break 
        except RequestException as e:
            context.log.error(f"Failed to request page: {e}")
            time.sleep(5 * attempt)
    return response


def count_non_empty_columns(sheet, row=2): # Default row is 2
    """Function to count the number of non-empty columns in the specified row."""
    max_col = sheet.max_column
    non_empty_cols = 0
    for col in range(1, max_col + 1):
        if sheet.cell(row=row, column=col).value:
            non_empty_cols += 1
    return non_empty_cols

# Function to find the last non-empty row in the sheet
def find_last_row(sheet):
    for row in range(sheet.max_row, 0, -1):
        if any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
            return row
    return 0

def transform_results(row):
    """Function to transform the results column.
    If the result is '...', return 'Відсутній'.
    """
    if row == '...':
        return 'Відсутній'
    else:
        return row # return unchanged result if not '...'

