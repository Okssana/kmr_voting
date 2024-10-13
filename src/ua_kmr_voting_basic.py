
from pprint import pprint
from urllib.parse import urljoin
from requests.exceptions import RequestException

from datetime import datetime
import lxml.html
import time
import os
import zipfile
from memorious.operations.extract import extract_zip
import pathlib
from openpyxl import load_workbook, Workbook
import hashlib
import json
import re 

zipfile_paths = []
CHUNK_SIZE = 100 
URL_FULL = "https://kmr.gov.ua/uk/result_golosuvanya?title=&field_start_date_n_h_value%5Bmin%5D=&field_start_date_n_h_value%5Bmax%5D=&page="
URL = 'https://kmr.gov.ua/uk/result_golosuvanya?title=&field_start_date_n_h_value[min]&field_start_date_n_h_value[max]&page=0&start=0'



HEADERS_XLSX = ["id",  "time", "question", "status", "short_name", "result"]



def request_page(context, url):

    for attempt in range(1, 5):
        try:
            response = context.http.get(url)
            break 
        except RequestException as e:
            context.log.error(f"Failed to request page: {e}")
            time.sleep(5 * attempt)
    return response


def count_non_empty_columns(sheet, row=2):
    """Function to count the number of non-empty columns in the specified row."""
    max_col = sheet.max_column
    non_empty_cols = 0
    for col in range(1, max_col + 1):
        if sheet.cell(row=row, column=col).value:
            non_empty_cols += 1
    return non_empty_cols

# Function to find the last non-empty row in the sheet
def find_last_row(sheet):
    for row in range(sheet.max_row, 0, -1):
        if any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
            return row
    return 0

def transform_results(row):
    """Function to transform the results column.
    If the result is '...', return 'Відсутній'.
    """
    if row == '...':
        return 'Відсутній'
    else:
        return row # return unchanged result if not '...'


def init(context, data):


    response_last_page = request_page(context, URL)
    doc_last_page = lxml.html.fromstring(response_last_page.text)

    ### Get the Last Page Number
    last_page = doc_last_page.xpath('//a[@class="field-content"][@title="До останньої сторінки"]/@href')
    last_page = last_page[0]
    last_page = last_page.split('&')[-2].split('=')[-1]
    last_page = int(last_page)

    # Page range for testing purposes: XLSX files are in this range.
    for i in range(16, 17): # Range should be (1, last_page + 1), but for testing purposes, it is set to (1, 2) or (1, 3) etc.
        context.log.info(f"Processing page: {i}")   

        URL_FULL_PAGE = URL_FULL + str(i)
        response_main_page = request_page(context, URL_FULL_PAGE)
        doc = lxml.html.fromstring(response_main_page.text)

        block = doc.xpath('//div[@class="view-content"]')

        #  "Результати поіменного голосування" are in block[0] only.

        for i in block[:1]: # change to [:1] to process only the first block

            for j in i.getchildren():
                # j is a separate block (row) with a ZIP or XLSX file: views-row views-row-{i} 
                print("J:", j.text_content(), "\n")

                if 'zip' in j.text_content(): # JSON files are stored in ZIP files

                    link = j.xpath('.//a/@href')[1]

                    # Download the zip file
                    response = request_page(context, link)
                    file_name = link.split('/')[-1]

                    zipfile_path = pathlib.Path(response.file_path)  
                    zipfile_paths.append(zipfile_path) 
                    extract_to = pathlib.Path(context.work_path)
                    context.log.info(f"Downloaded and stored ZIP file: {file_name}")

                    # Extract the ZIP file 
                    with zipfile.ZipFile(zipfile_path, "r") as zf:
                        zf.extractall(extract_to)

                    context.log.info(f"ZIP file {zipfile_path} was extracted to {extract_to}")

                    # read JSON files in the extracted folder
                    #process_extracted_files(context, extract_to)

                    # with open(file_name, 'wb') as f:
                    #     f.write(response.text)
                    
                    # extract the file
                    # extract_zip(context, file_name)
                    # os.remove(file_name)
                    # break

                ## Process XLSX files in the block ##
                elif 'xlsx' in j.text_content():
                    link_xlsx = j.xpath('.//a/@href')[1]
                    print("Link XLSX:", link_xlsx)
                    context.log.info(f"Found XLSX file: {link_xlsx}")
                    # download the file
                    response = request_page(context, link_xlsx)

                    if response:
                        path = pathlib.Path(response.file_path)
                        print("Path:", path)

                        file_name = link_xlsx.split('/')[-1]

                        print("File path MIS:", file_name)

                        # Read .xlsx file from path
                        with open(path, 'rb') as f:
                            file_content = f.read()
                            
                        new_path = path.with_suffix('.xlsx')
                        print("New path:", new_path)

                        # Rewrite the file with .xlsx extension
                        with open(new_path, 'wb') as f:
                            f.write(file_content)

                        # Process a XLSX file
                        process_xlsx(context, new_path, file_name, URL_FULL_PAGE)
                        
                        context.log.info(f"Processed XLSX file: {file_name}")

                    else:
                        context.log.info(f"Skipping file:")



def process_xlsx(context, data, file_name, URL_FULL_PAGE):
    # Create a new workbook for the combined long table
    combined_wb = Workbook()
    combined_sheet = combined_wb.active
    combined_sheet.append(HEADERS_XLSX)

    xlsx_files = [file for file in os.listdir(context.work_path) if file.endswith('.xlsx')] # Get all XLSX files in the work path

    print("XLSX files:", xlsx_files)
    context.log.info(f" Number of XLSX files: {len(xlsx_files)}")

    for xlsx_file in xlsx_files:

        context.log.info(f"Processing file: {xlsx_file}")
        file_path = os.path.join(context.work_path, xlsx_file)

        wb = load_workbook(file_path, read_only=True) # read_only=True to avoid memory issues
        ws = wb[wb.sheetnames[0]]

        sheet = wb[wb.sheetnames[0]] # Select the first sheet
        
        max_row = find_last_row(sheet) 
        non_empty_col_count = count_non_empty_columns(sheet)

        # Determine the starting column based on the number of columns
        start_col = 5 if non_empty_col_count == 125 else 6 if non_empty_col_count == 126 else 5 # 5 or 6

        fixed_cols_range = range(2, start_col) if non_empty_col_count == 126 else range(1, start_col) if non_empty_col_count == 125 else range(1, start_col)

        # Iterate over rows in chunks
        for row in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
            fixed_data = [row[col - 1] for col in fixed_cols_range]

            for col in range(start_col, start_col + 121):
                entity_name = ws.cell(row=1, column=col).value
                result = row[col - 1]  # Adjust column for 0-indexed
                new_row = fixed_data + [entity_name, result]
                combined_sheet.append(new_row)
        
        # Save and emit
        date_time = datetime.now().strftime("%Y-%m-%d_%H-%M")

        # Save the combined XLSX file
        combined_path = f"/Users/Oksana/Documents/PERSONAL_PRJCTS/ua_kmr_voting/excel/long_table_{date_time}_{file_name}"

        combined_wb.save(combined_path)
        context.log.info(f"Combined XLSX file saved: {combined_path}")

        
        emit_rows_from_excel(context, combined_path, xlsx_file, file_name, URL_FULL_PAGE)

         
def emit_rows_from_excel(context, file_path, xlsx_file, file_name, URL_FULL_PAGE):

    """
    Function to emit rows from an Excel file.
    
    Args:
        file_path (str): The path to the Excel file.
        file_name (str): The name of the transformed file. Extension is added.
        file_name (str): The link to the Excel file. Original source and file name.
        URL_FULL_PAGE (str): The URL of the full page.
    """

    wb = load_workbook(file_path,  read_only=True)
    sheet = wb.active
    
    # Iterate over rows starting from the second row (assuming the first row is headers)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data_rows = {
            "id": row[0],
            "time": row[1],
            "question": row[2],
            "status": row[3],
            "short_name": row[4],
            "result": transform_results(row[5]), # "If the result is '...', return 'Відсутній'."
            "file_name": xlsx_file,
            "file_name_original": file_name,
            "source_url": URL_FULL_PAGE
        }
        context.emit(data=data_rows)

# Counter to ensure uniqueness
counter = 0


def store_exel(context, data):
    global counter
    counter += 1

    context.log.info(f"Storing data excel for the day : {data.get('time')} and ID Question: {data.get('id')}")

    unique_string = str(int(data.get('id')) + int(counter))
    unique_id = hashlib.md5(unique_string.encode()).hexdigest()
    time.sleep(1)

    # Ensure all fields match the expected types in your database

    doc_data = {
        #'id': unique_id,  # Ensure id is always a string
        'id_original': data.get('id'), # Keep original ID for reference
        "unique_id": unique_id,
        'time': str(data.get('time')), 
        'question': data.get('question') or "", 
        'status': data.get('status') or "",  
        'short_name': data.get('short_name') or "",  
        'result': data.get('result') or "", 
        'source_file': data.get('file_name') or None,  
        "source_file_original": data.get('file_name_original') or None,
        'retrieved_at': datetime.now().isoformat(),
        'source_url': data.get('source_url') or None  
    }

    table = context.datastore['ua_kmr_voting_xlsx']
    time.sleep(1)
    table.upsert(doc_data, [#'id', # id is present by default, no need to specify
                            'id_original', 'unique_id',
                            'time', 'question', 'status', 'short_name', 'result',
                            'source_file', 'source_file_original',
                            'retrieved_at', 'source_url'])

    context.log.info(f"Stored Number of Question in data dict: {data.get('id')} for the day : {data.get('time')}")



# Function to preprocess JSON data
def preprocess_json(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Correct invalid newline characters within strings
    corrected_content = re.sub(r'(?<!\\)"\s*\n\s*"', r'\"', content)
    
    return corrected_content

def process_extracted_files(context, extract_to):
    json_files = [file for file in os.listdir(extract_to) if file.endswith('.json')]
    context.log.info(f"Number of JSON files: {len(json_files)} - {json_files[:1]} ")

    for json_file in json_files[:5]:
        context.log.info(f"Processing file: {json_file} with a length of {len(json_file)}")

        file_path = os.path.join(extract_to, json_file)

        # Preprocess and read the JSON file
        corrected_json = preprocess_json(file_path)

        # Parse the JSON data
        try:
            json_data = json.loads(corrected_json)
        except json.JSONDecodeError as e:
            print(f"Error decoding JSON: {e}")
            return

        # Extract and store data in the dict
        dict_d = {
            'file_name': json_file,
            'json_data': json_data,
            'file_path': file_path,
            'DocTime': json_data['DocTime'],

            'source_file': json_file,
            'orgName': json_data['OrgName'],
            'SName': json_data['SName'],
            'GLType': json_data['GLType'],
            'GLTime': json_data['GLTime'],
            'PD_NPP': json_data['PD_NPP'],
            'GL_Text': json_data['GL_Text'],
            'DPList': json_data['DPList']
        }
        print("*******************")
        context.emit(data=dict_d)
        context.log.info(f"Extracted data emitted: {json_file}")


# def store(context, data):
#     # Extract main document data
#     context.log.info(f"Storing data: {data.get('file_name')}")
#     doc_data = {
#         'id': data.get('id'),
#         'file_name': data.get('file_name'),

#         'source_url': data.get('source_url'),
#         'orgName': data.get('OrgName'),
#         'SName': data.get('SName'),
#         'GLType': data.get('GLType'),
#         'GLTime': data.get('GLTime'),
#         'PD_NPP': data.get('PD_NPP'),
#         'GL_Text': data.get('GL_Text'),
#         'DocTime': data.get('DocTime')
#     }

#     # Process and combine DPList entries with main document data
#     combined_data = []
#     dp_list = data.get('DPList', [])
#     for dp_entry in dp_list:
#         combined_entry = doc_data.copy()
#         combined_entry.update({
#             'DPName': dp_entry.get('DPName'),
#             'DPGolos': 'Відсутній' if dp_entry.get('DPGolos') == '.........' else dp_entry.get('DPGolos')
#         })
#         #print("Combined entry:", combined_entry)
#         combined_data.append(combined_entry)
    
#     if not dp_list:
#         combined_data.append(doc_data)
    
#     # Insert combined data into a single table
#     table_name = context.params.get('dataset')
#     table = context.datastore[table_name]
#     for entry in combined_data:
#         #print("Entry:", entry, "\n")
#         table.upsert(entry, ['id', 'file_name', 
#                              'SName', 'GLType', 'GLTime', 'PD_NPP', 'GL_Text', 
#                              'DocTime', 'DPName', 'DPGolos'])


